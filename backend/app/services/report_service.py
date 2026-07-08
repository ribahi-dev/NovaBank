"""Génération de rapports serveur en PDF et Excel (CdC Module 10).

Deux formats à partir des mêmes données agrégées :
  - PDF (reportlab) : document imprimable pour la direction ;
  - Excel (openpyxl) : fichier exploitable pour une analyse plus poussée.

Les fonctions renvoient des `bytes` : le router les sert en téléchargement.
Aucune logique HTTP ici (couche service).
"""

import io
from datetime import datetime, timezone
from decimal import Decimal

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import Account, Alert, Client, RiskScore, Transaction

BRAND_ORANGE = "#F08100"
DARK = "#1F2427"


def _collect_kpis(db: Session) -> dict:
    """Rassemble les indicateurs clés (mêmes chiffres que le dashboard)."""
    deposits = db.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            Transaction.transaction_type == "deposit"
        )
    )
    withdrawals = db.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            Transaction.transaction_type == "withdrawal"
        )
    )
    return {
        "clients": db.scalar(select(func.count()).select_from(Client).where(Client.is_active.is_(True))),
        "accounts": db.scalar(select(func.count()).select_from(Account)),
        "transactions": db.scalar(select(func.count()).select_from(Transaction)),
        "open_alerts": db.scalar(select(func.count()).select_from(Alert).where(Alert.status == "open")),
        "deposits": deposits or Decimal("0"),
        "withdrawals": withdrawals or Decimal("0"),
        "avg_risk": db.scalar(select(func.avg(RiskScore.score))),
    }


def _recent_alerts(db: Session, limit: int = 15):
    return db.scalars(
        select(Alert).order_by(Alert.created_at.desc()).limit(limit)
    ).all()


def generate_activity_pdf(db: Session) -> bytes:
    """Rapport d'activité au format PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    kpis = _collect_kpis(db)
    alerts = _recent_alerts(db)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor(BRAND_ORANGE))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor(DARK))
    normal = styles["Normal"]

    generated = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    story = [
        Paragraph("NovaBank — Rapport d'activité de l'agence", title),
        Paragraph(f"Généré le {generated}", normal),
        Spacer(1, 0.6 * cm),
        Paragraph("Indicateurs clés", h2),
    ]

    kpi_rows = [
        ["Indicateur", "Valeur"],
        ["Clients actifs", str(kpis["clients"])],
        ["Comptes", str(kpis["accounts"])],
        ["Transactions", str(kpis["transactions"])],
        ["Total des dépôts (MAD)", f"{kpis['deposits']:,.2f}"],
        ["Total des retraits (MAD)", f"{kpis['withdrawals']:,.2f}"],
        ["Alertes ouvertes", str(kpis["open_alerts"])],
        ["Score de risque moyen", f"{kpis['avg_risk']:.1f}/100" if kpis["avg_risk"] else "—"],
    ]
    kpi_table = Table(kpi_rows, colWidths=[9 * cm, 6 * cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_ORANGE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E7EC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story += [kpi_table, Spacer(1, 0.8 * cm), Paragraph("Dernières alertes", h2)]

    alert_rows = [["Date", "Niveau", "Statut", "Message"]]
    for a in alerts:
        alert_rows.append([
            a.created_at.strftime("%d/%m/%Y %H:%M"),
            a.level,
            a.status,
            (a.message[:70] + "…") if len(a.message) > 70 else a.message,
        ])
    if len(alert_rows) == 1:
        alert_rows.append(["—", "—", "—", "Aucune alerte"])
    alert_table = Table(alert_rows, colWidths=[3.2 * cm, 2 * cm, 2 * cm, 7.8 * cm])
    alert_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(DARK)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E4E7EC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story += [
        alert_table,
        Spacer(1, 1 * cm),
        Paragraph(
            "Document généré automatiquement par la plateforme NovaBank — données simulées.",
            ParagraphStyle("f", parent=normal, fontSize=7, textColor=colors.grey),
        ),
    ]

    doc.build(story)
    return buffer.getvalue()


def generate_transactions_xlsx(db: Session, limit: int = 500) -> bytes:
    """Export Excel des transactions avec leur score de risque."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = db.scalars(
        select(Transaction).order_by(Transaction.created_at.desc()).limit(limit)
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["ID", "Date", "Type", "Montant (MAD)", "Ville", "Score de risque", "Moteur", "Explication"]
    ws.append(headers)
    header_fill = PatternFill(start_color="F08100", end_color="F08100", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for t in rows:
        score = t.risk_score
        ws.append([
            t.id,
            t.created_at.strftime("%d/%m/%Y %H:%M"),
            t.transaction_type,
            float(t.amount),
            t.city or "",
            score.score if score else "",
            score.model_version if score else "",
            score.explanation if score else "",
        ])

    # Largeurs de colonnes lisibles.
    widths = [8, 18, 12, 15, 14, 14, 14, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
